"""
Finance views — dashboard, transactions, budgets, sponsorships, shop
"""
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_POST

from core.decorators import manager_required, staff_required
from .models import (
    Budget, BudgetCategory, Transaction, Sponsorship, Investment,
    Product, Order, OrderItem, Cart, CartItem,
)
from .forms import (
    BudgetForm, TransactionForm, SponsorshipForm, InvestmentForm,
    ProductForm, OrderForm,
)


@manager_required
def finance_dashboard(request):
    today = timezone.now().date()
    month_start = today.replace(day=1)
    year_start  = today.replace(month=1, day=1)
    club = getattr(request, "active_club", None)

    tx_qs      = Transaction.objects.all()
    sponsor_qs = Sponsorship.objects.all()
    invest_qs  = Investment.objects.all()
    budget_qs  = Budget.objects.select_related("category")

    if club:
        tx_qs      = tx_qs.filter(team__club=club)
        sponsor_qs = sponsor_qs.filter(club=club)
        invest_qs  = invest_qs.filter(club=club)
        budget_qs  = budget_qs.filter(team__club=club)

    # Summary KPIs
    monthly_income  = tx_qs.filter(
        transaction_type=Transaction.Type.INCOME, date__gte=month_start, status=Transaction.Status.COMPLETED,
    ).aggregate(t=Sum("amount"))["t"] or 0
    monthly_expense = tx_qs.filter(
        transaction_type=Transaction.Type.EXPENSE, date__gte=month_start, status=Transaction.Status.COMPLETED,
    ).aggregate(t=Sum("amount"))["t"] or 0
    annual_income   = tx_qs.filter(
        transaction_type=Transaction.Type.INCOME, date__gte=year_start,
    ).aggregate(t=Sum("amount"))["t"] or 0
    total_sponsorship = sponsor_qs.filter(
        status=Sponsorship.Status.ACTIVE
    ).aggregate(t=Sum("amount"))["t"] or 0
    total_investment = invest_qs.filter(
        is_active=True
    ).aggregate(t=Sum("amount"))["t"] or 0

    recent_transactions = tx_qs.select_related("category", "team").order_by("-date")[:10]
    active_sponsorships = sponsor_qs.filter(status=Sponsorship.Status.ACTIVE)[:6]
    budgets = budget_qs[:6]

    # Category breakdown for pie chart
    expense_by_category = (
        tx_qs
        .filter(transaction_type=Transaction.Type.EXPENSE, date__gte=month_start)
        .values("category__name", "category__color")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )

    return render(request, "finance/dashboard.html", {
        "monthly_income":   monthly_income,
        "monthly_expense":  monthly_expense,
        "annual_income":    annual_income,
        "net_monthly":      monthly_income - monthly_expense,
        "total_sponsorship": total_sponsorship,
        "total_investment":  total_investment,
        "recent_transactions": recent_transactions,
        "active_sponsorships": active_sponsorships,
        "budgets": budgets,
        "expense_by_category": list(expense_by_category),
    })


@manager_required
def transaction_list(request):  # noqa
    qs = Transaction.objects.select_related("category", "team", "created_by").order_by("-date")
    club = getattr(request, "active_club", None)
    if club:
        qs = qs.filter(team__club=club)
    tx_type  = request.GET.get("type")
    status   = request.GET.get("status")
    category = request.GET.get("category")
    q        = request.GET.get("q", "")
    date_from = request.GET.get("date_from")
    date_to   = request.GET.get("date_to")

    if tx_type:
        qs = qs.filter(transaction_type=tx_type)
    if status:
        qs = qs.filter(status=status)
    if category:
        qs = qs.filter(category_id=category)
    if q:
        qs = qs.filter(title__icontains=q)
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    total_income  = qs.filter(transaction_type=Transaction.Type.INCOME).aggregate(t=Sum("amount"))["t"] or 0
    total_expense = qs.filter(transaction_type=Transaction.Type.EXPENSE).aggregate(t=Sum("amount"))["t"] or 0

    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get("page", 1))

    categories = BudgetCategory.objects.all()
    return render(request, "finance/transactions.html", {
        "page": page,
        "total_income":  total_income,
        "total_expense": total_expense,
        "net": total_income - total_expense,
        "categories": categories,
        "type_choices": Transaction.Type.choices,
        "status_choices": Transaction.Status.choices,
    })


@manager_required
def transaction_create(request):
    form = TransactionForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        tx = form.save(commit=False)
        tx.created_by = request.user
        tx.save()
        from core.models import ActivityLog
        ActivityLog.log(request.user, ActivityLog.Action.CREATE,
                        f"Added transaction: {tx.title} €{tx.amount}", "Transaction", tx.pk)
        messages.success(request, _("Transaction added."))
        if request.htmx:
            return HttpResponse(
                '<div class="text-emerald-600">✓ Transaction saved</div>',
                headers={"HX-Trigger": "transactionAdded"},
            )
        return redirect("finance:transactions")
    return render(request, "finance/transaction_form.html", {"form": form, "title": _("New Transaction")})


@manager_required
def transaction_edit(request, pk):
    tx = get_object_or_404(Transaction, pk=pk)
    form = TransactionForm(request.POST or None, request.FILES or None, instance=tx)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Transaction updated."))
        return redirect("finance:transactions")
    return render(request, "finance/transaction_form.html", {
        "form": form, "title": _("Edit Transaction"), "transaction": tx,
    })


@manager_required
def budget_list(request):
    budgets = Budget.objects.select_related("category", "season", "team")
    club = getattr(request, "active_club", None)
    if club:
        budgets = budgets.filter(team__club=club)
    return render(request, "finance/budgets.html", {"budgets": budgets})


@manager_required
def budget_create(request):
    form = BudgetForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        budget = form.save(commit=False)
        budget.created_by = request.user
        budget.save()
        messages.success(request, _("Budget created."))
        return redirect("finance:budgets")
    return render(request, "finance/budget_form.html", {"form": form})


@manager_required
def sponsorship_list(request):
    sponsors = Sponsorship.objects.all().order_by("-start_date")
    club = getattr(request, "active_club", None)
    if club:
        sponsors = sponsors.filter(club=club)
    return render(request, "finance/sponsorships.html", {"sponsorships": sponsors})


@manager_required
def sponsorship_create(request):
    form = SponsorshipForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        sponsor = form.save(commit=False)
        club = getattr(request, "active_club", None)
        if club:
            sponsor.club = club
        sponsor.save()
        messages.success(request, _("Sponsorship added."))
        return redirect("finance:sponsorships")
    return render(request, "finance/sponsorship_form.html", {"form": form})


@manager_required
def investment_list(request):
    investments = Investment.objects.order_by("-date")
    club = getattr(request, "active_club", None)
    if club:
        investments = investments.filter(club=club)
    total = investments.filter(is_active=True).aggregate(t=Sum("amount"))["t"] or 0
    return render(request, "finance/investments.html", {"investments": investments, "total": total})


@manager_required
def investment_create(request):
    form = InvestmentForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        investment = form.save(commit=False)
        club = getattr(request, "active_club", None)
        if club:
            investment.club = club
        investment.save()
        messages.success(request, _("Investment recorded."))
        return redirect("finance:investments")
    return render(request, "finance/investment_form.html", {"form": form})


# ─── Shop ────────────────────────────────────────────────────────────────────

@login_required
def shop(request):
    products = Product.objects.filter(is_active=True)
    category = request.GET.get("category")
    if category:
        products = products.filter(category=category)
    cart = _get_or_create_cart(request)
    return render(request, "finance/shop.html", {
        "products": products,
        "category_choices": Product.Category.choices,
        "current_category": category,
        "cart": cart,
    })


@login_required
def product_detail(request, slug):
    product = get_object_or_404(Product, slug=slug, is_active=True)
    cart = _get_or_create_cart(request)
    return render(request, "finance/product_detail.html", {"product": product, "cart": cart})


@login_required
@require_POST
def cart_add(request, product_id):
    product = get_object_or_404(Product, pk=product_id, is_active=True)
    # Ensure session exists for anonymous users
    if not request.session.session_key:
        request.session.create()
    cart    = _get_or_create_cart(request)
    qty     = int(request.POST.get("quantity", 1))

    item, created = CartItem.objects.get_or_create(
        cart=cart, product=product, defaults={"price": product.price, "quantity": qty},
    )
    if not created:
        item.quantity += qty
        item.save(update_fields=["quantity"])

    if request.htmx:
        return render(request, "partials/cart_badge.html", {"cart": cart})
    messages.success(request, _("%(name)s added to cart.") % {"name": product.name})
    return redirect("finance:cart")


@login_required
@require_POST
def cart_remove(request, item_id):
    cart = _get_or_create_cart(request)
    item = get_object_or_404(CartItem, pk=item_id, cart=cart)
    item.delete()
    cart = _get_or_create_cart(request)
    if request.htmx:
        return render(request, "finance/partials/cart_items.html", {"cart": cart})
    return redirect("finance:cart")


@login_required
def cart_view(request):
    cart = _get_or_create_cart(request)
    return render(request, "finance/cart.html", {"cart": cart})


@login_required
def checkout(request):
    """Checkout page — shows shipping form, then redirects to Stripe."""
    cart = _get_or_create_cart(request)
    if cart.item_count == 0:
        messages.warning(request, _("Your cart is empty."))
        return redirect("finance:shop")
    return render(request, "finance/checkout.html", {"cart": cart})


@login_required
@require_POST
def checkout_submit(request):
    """Create pending Order then redirect to Stripe Checkout Session."""
    import stripe
    from django.conf import settings as dj_settings

    cart = _get_or_create_cart(request)
    if cart.item_count == 0:
        messages.warning(request, _("Your cart is empty."))
        return redirect("finance:shop")

    # Build the order first (pending payment)
    order = Order.objects.create(
        customer=request.user,
        status=Order.Status.PENDING,
        shipping_address=request.POST.get("shipping_address", ""),
        notes=request.POST.get("notes", ""),
    )
    for item in cart.cart_items.select_related("product"):
        OrderItem.objects.create(
            order=order, product=item.product,
            quantity=item.quantity, price=item.price,
        )
    order.calculate_total()

    # Build Stripe line items
    stripe.api_key = dj_settings.STRIPE_SECRET_KEY
    base_url = request.build_absolute_uri("/").rstrip("/")

    line_items = []
    for item in order.items.select_related("product"):
        line_items.append({
            "price_data": {
                "currency": dj_settings.STRIPE_CURRENCY,
                "unit_amount": int(item.price * 100),  # cents
                "product_data": {
                    "name": item.product.name,
                    "images": (
                        [request.build_absolute_uri(item.product.image.url)]
                        if item.product.image else []
                    ),
                },
            },
            "quantity": item.quantity,
        })

    try:
        session = stripe.checkout.Session.create(
            payment_method_types=["card"],
            line_items=line_items,
            mode="payment",
            customer_email=request.user.email,
            metadata={"order_id": order.pk},
            success_url=f"{base_url}/finance/orders/{order.pk}/payment/success/",
            cancel_url=f"{base_url}/finance/orders/{order.pk}/payment/cancel/",
        )
        # Store session ID on order for webhook verification
        order.notes = (order.notes + f"\nstripe_session:{session.id}").strip()
        order.save(update_fields=["notes"])
        # Clear cart
        cart.cart_items.all().delete()
        return redirect(session.url, permanent=False)
    except stripe.error.StripeError as e:
        order.delete()
        messages.error(request, _("Payment error: %(err)s") % {"err": str(e)})
        return redirect("finance:checkout")


@login_required
def payment_success(request, pk):
    """Stripe redirects here after successful payment."""
    order = get_object_or_404(Order, pk=pk, customer=request.user)
    if order.status == Order.Status.PENDING:
        order.status = Order.Status.CONFIRMED
        order.save(update_fields=["status"])
        # Record as income transaction
        Transaction.objects.create(
            title=_("Shop Order #%(pk)s") % {"pk": order.pk},
            transaction_type=Transaction.Type.INCOME,
            amount=order.total,
            date=timezone.now().date(),
            status=Transaction.Status.COMPLETED,
            created_by=request.user,
        )
    messages.success(request, _("Payment successful! Order #%(pk)s confirmed.") % {"pk": order.pk})
    return render(request, "finance/payment_success.html", {"order": order})


@login_required
def payment_cancel(request, pk):
    """Stripe redirects here when the user cancels payment."""
    order = get_object_or_404(Order, pk=pk, customer=request.user)
    if order.status == Order.Status.PENDING:
        order.status = Order.Status.CANCELLED
        order.save(update_fields=["status"])
    messages.warning(request, _("Payment was cancelled. Your order has been voided."))
    return redirect("finance:cart")


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def stripe_webhook(request):
    """Stripe webhook — handles payment_intent.succeeded as a backup."""
    import stripe
    from django.conf import settings as dj_settings

    stripe.api_key = dj_settings.STRIPE_SECRET_KEY
    payload   = request.body
    sig_header = request.META.get("HTTP_STRIPE_SIGNATURE", "")

    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, dj_settings.STRIPE_WEBHOOK_SECRET
        )
    except (ValueError, stripe.error.SignatureVerificationError):
        return HttpResponse(status=400)

    if event["type"] == "checkout.session.completed":
        session  = event["data"]["object"]
        order_id = session.get("metadata", {}).get("order_id")
        if order_id:
            try:
                order = Order.objects.get(pk=order_id)
                if order.status == Order.Status.PENDING:
                    order.status = Order.Status.CONFIRMED
                    order.save(update_fields=["status"])
            except Order.DoesNotExist:
                pass

    return HttpResponse(status=200)


@login_required
def order_list(request):
    if request.user.can_manage:
        qs = Order.objects.all().order_by("-created_at")
    else:
        qs = Order.objects.filter(customer=request.user).order_by("-created_at")
    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get("page", 1))
    return render(request, "finance/orders.html", {"page": page})


@login_required
def order_detail(request, pk):
    order = get_object_or_404(Order.objects.prefetch_related("items__product"), pk=pk)
    return render(request, "finance/order_detail.html", {
        "order": order, "status_choices": Order.Status.choices,
    })


@login_required
def order_invoice(request, pk):
    """
    Render a clean printable invoice for an order.
    Add ?format=pdf to get a downloadable PDF via ReportLab.
    """
    order = get_object_or_404(
        Order.objects.prefetch_related("items__product"),
        pk=pk,
    )
    # Only the customer or managers can view the invoice
    if not request.user.can_manage and order.customer != request.user:
        from django.http import Http404
        raise Http404

    invoice_number = f"INV-{order.created_at.strftime('%Y%m')}-{order.pk:04d}"

    if request.GET.get("format") == "pdf":
        return _render_invoice_pdf(order, invoice_number)

    return render(request, "finance/invoice.html", {
        "order": order,
        "invoice_number": invoice_number,
    })


def _render_invoice_pdf(order, invoice_number):
    """Generate a PDF invoice using ReportLab."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=20 * mm, leftMargin=20 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    primary = colors.HexColor("#059669")   # emerald-600
    dark    = colors.HexColor("#111827")   # gray-900
    muted   = colors.HexColor("#6B7280")   # gray-500
    light   = colors.HexColor("#F3F4F6")   # gray-100

    h1  = ParagraphStyle("h1",  fontSize=24, fontName="Helvetica-Bold", textColor=dark, leading=30)
    h2  = ParagraphStyle("h2",  fontSize=12, fontName="Helvetica-Bold", textColor=dark, leading=16)
    sub = ParagraphStyle("sub", fontSize=9,  fontName="Helvetica",      textColor=muted, leading=13)
    body = ParagraphStyle("body", fontSize=10, fontName="Helvetica",    textColor=dark,  leading=14)
    right = ParagraphStyle("right", fontSize=10, fontName="Helvetica-Bold", textColor=primary, alignment=TA_RIGHT)
    center = ParagraphStyle("center", fontSize=8, fontName="Helvetica", textColor=muted, alignment=TA_CENTER)

    story = []

    # ── Header ──────────────────────────────────────────────────────────────
    header_data = [
        [
            Paragraph("<b>ApexForge</b>", ParagraphStyle(
                "brand", fontSize=20, fontName="Helvetica-Bold", textColor=primary,
            )),
            Paragraph(f"<b>INVOICE</b>", ParagraphStyle(
                "inv", fontSize=20, fontName="Helvetica-Bold", textColor=dark, alignment=TA_RIGHT,
            )),
        ]
    ]
    header_table = Table(header_data, colWidths=["50%", "50%"])
    header_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(header_table)
    story.append(Spacer(1, 4 * mm))
    story.append(HRFlowable(width="100%", thickness=2, color=primary))
    story.append(Spacer(1, 5 * mm))

    # ── Invoice meta + customer block ────────────────────────────────────────
    customer_name = order.customer.get_full_name() or order.customer.email
    meta_data = [
        [
            Paragraph(f"Invoice Number:<br/><b>{invoice_number}</b>", body),
            Paragraph(
                f"Billed To:<br/><b>{customer_name}</b><br/>{order.customer.email}",
                body,
            ),
        ],
        [
            Paragraph(f"Date Issued:<br/><b>{order.created_at.strftime('%d %B %Y')}</b>", body),
            Paragraph(
                f"Shipping Address:<br/>{order.shipping_address or '—'}".replace("\n", "<br/>"),
                body,
            ),
        ],
        [
            Paragraph(f"Status:<br/><b>{order.get_status_display()}</b>", body),
            Paragraph("", body),
        ],
    ]
    meta_table = Table(meta_data, colWidths=["50%", "50%"], rowHeights=None)
    meta_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 6 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 5 * mm))

    # ── Line items table ─────────────────────────────────────────────────────
    story.append(Paragraph("Order Items", h2))
    story.append(Spacer(1, 3 * mm))

    table_head = [["Product", "Unit Price", "Qty", "Subtotal"]]
    table_rows = []
    for item in order.items.all():
        table_rows.append([
            Paragraph(item.product.name if item.product else "—", body),
            f"€{item.price:.2f}",
            str(item.quantity),
            f"€{item.subtotal:.2f}",
        ])

    # Subtotal / shipping / total footer rows
    table_rows.append(["", "", "Subtotal", f"€{order.total:.2f}"])
    table_rows.append(["", "", "Shipping", "Free"])
    table_rows.append(["", "", "TOTAL", f"€{order.total:.2f}"])

    all_rows = table_head + table_rows
    items_table = Table(
        all_rows,
        colWidths=["50%", "18%", "12%", "20%"],
    )
    n = len(table_rows)
    last = len(all_rows) - 1
    items_table.setStyle(TableStyle([
        # Header
        ("BACKGROUND",  (0, 0), (-1, 0), primary),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("TOPPADDING",  (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        # Body rows
        ("FONTSIZE",    (0, 1), (-1, -1), 10),
        ("TOPPADDING",  (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, last - 3), [colors.white, light]),
        # Align amounts right
        ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
        # Summary rows styling
        ("FONTNAME",    (2, last - 1), (-1, last - 1), "Helvetica-Bold"),
        ("BACKGROUND",  (0, last), (-1, last), dark),
        ("TEXTCOLOR",   (0, last), (-1, last), colors.white),
        ("FONTNAME",    (0, last), (-1, last), "Helvetica-Bold"),
        ("FONTSIZE",    (0, last), (-1, last), 11),
        ("TOPPADDING",  (0, last), (-1, last), 7),
        ("BOTTOMPADDING", (0, last), (-1, last), 7),
        ("LINEABOVE",   (0, last - 2), (-1, last - 2), 0.5, colors.HexColor("#E5E7EB")),
        ("GRID",        (0, 0), (-1, 0), 0, colors.transparent),
        ("BOX",         (0, 0), (-1, last), 0, colors.transparent),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 10 * mm))

    # ── Footer ───────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(
        "Thank you for your purchase! · ApexForge Sports Platform · support@apexforge.com",
        center,
    ))

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="invoice-{invoice_number}.pdf"'
    )
    return response


@manager_required
def export_transactions_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["ID", "Date", "Title", "Type", "Amount", "Category", "Status"]
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="059669")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf
        c.fill = hfill

    tx_export = Transaction.objects.select_related("category").order_by("-date")
    club = getattr(request, "active_club", None)
    if club:
        tx_export = tx_export.filter(team__club=club)
    for row, tx in enumerate(tx_export, 2):
        ws.cell(row=row, column=1, value=tx.pk)
        ws.cell(row=row, column=2, value=str(tx.date))
        ws.cell(row=row, column=3, value=tx.title)
        ws.cell(row=row, column=4, value=tx.get_transaction_type_display())
        ws.cell(row=row, column=5, value=float(tx.amount))
        ws.cell(row=row, column=6, value=tx.category.name if tx.category else "")
        ws.cell(row=row, column=7, value=tx.get_status_display())

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=transactions.xlsx"
    wb.save(response)
    return response


def _get_or_create_cart(request):
    if request.user.is_authenticated:
        cart = Cart.objects.filter(user=request.user).first()
        if cart is None:
            cart = Cart.objects.create(user=request.user)
    else:
        key = request.session.session_key or ""
        cart = Cart.objects.filter(session_key=key).first()
        if cart is None:
            cart = Cart.objects.create(session_key=key)
    return cart


@manager_required
def budget_edit(request, pk):
    budget = get_object_or_404(Budget, pk=pk)
    form = BudgetForm(request.POST or None, instance=budget)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Budget updated."))
        return redirect("finance:budgets")
    return render(request, "finance/budget_form.html", {"form": form, "budget": budget})


@manager_required
def investment_edit(request, pk):
    investment = get_object_or_404(Investment, pk=pk)
    form = InvestmentForm(request.POST or None, instance=investment)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Investment updated."))
        return redirect("finance:investments")
    return render(request, "finance/investment_form.html", {"form": form, "investment": investment})


@manager_required
def sponsorship_edit(request, pk):
    sponsorship = get_object_or_404(Sponsorship, pk=pk)
    form = SponsorshipForm(request.POST or None, request.FILES or None, instance=sponsorship)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Sponsorship updated."))
        return redirect("finance:sponsorships")
    return render(request, "finance/sponsorship_form.html", {"form": form, "sponsorship": sponsorship})


@manager_required
def product_list(request):
    """Staff view: manage all shop products."""
    products = Product.objects.all().order_by("category", "name")
    return render(request, "finance/product_list.html", {"products": products})


@manager_required
def product_create(request):
    form = ProductForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Product added to store."))
        return redirect("finance:product_list")
    return render(request, "finance/product_form.html", {"form": form, "title": _("Add Product")})


@manager_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    form = ProductForm(request.POST or None, request.FILES or None, instance=product)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Product updated."))
        return redirect("finance:product_detail", slug=product.slug)
    return render(request, "finance/product_form.html", {"form": form, "product": product})


@manager_required
@require_POST
def order_status(request, pk):
    order = get_object_or_404(Order, pk=pk)
    status = request.POST.get("status")
    if status and status in dict(Order.Status.choices):
        order.status = status
        order.save(update_fields=["status"])
        messages.success(request, _("Order status updated."))
    return redirect("finance:order_detail", pk=pk)
