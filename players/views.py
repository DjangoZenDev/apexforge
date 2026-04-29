"""
Players views — profiles, stats, injuries, performance
"""
import io
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_POST

from .models import Player, PlayerStats, InjuryLog, PerformanceMetric
from .forms import PlayerForm, PlayerStatsForm, InjuryLogForm, PerformanceMetricForm
from core.pdf import player_profile_pdf


@login_required
def player_list(request):
    qs = Player.objects.prefetch_related("rosters__team")
    # Filter by active club via roster → team → club
    club = getattr(request, "active_club", None)
    if club:
        qs = qs.filter(rosters__team__club=club).distinct()
    q = request.GET.get("q", "")
    status = request.GET.get("status")
    position = request.GET.get("position")
    sport = request.GET.get("sport")

    if q:
        qs = qs.filter(full_name__icontains=q)
    if status:
        qs = qs.filter(status=status)
    if position:
        qs = qs.filter(position__icontains=position)
    if sport:
        qs = qs.filter(sport=sport)

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get("page", 1))

    if request.htmx:
        return render(request, "players/partials/player_rows.html", {"page": page})

    return render(request, "players/list.html", {
        "page": page, "query": q,
        "status_choices": Player.Status.choices,
        "current_status": status,
    })


@login_required
def player_detail(request, pk):
    player = get_object_or_404(Player, pk=pk)
    stats   = PlayerStats.objects.filter(player=player).select_related("team", "season")
    injuries = InjuryLog.objects.filter(player=player)
    metrics  = PerformanceMetric.objects.filter(player=player).order_by("-date")[:20]
    rosters  = player.rosters.select_related("team", "season").order_by("-joined_date")

    # Chart data — last 10 metrics
    metric_labels = [str(m.date) for m in metrics][::-1]
    metric_values = [float(m.value) for m in metrics][::-1]

    from scouting.models import ScoutReport
    scout_reports = ScoutReport.objects.filter(player=player).select_related("scout")

    return render(request, "players/detail.html", {
        "player": player,
        "stats": stats,
        "injuries": injuries,
        "metrics": metrics,
        "rosters": rosters,
        "scout_reports": scout_reports,
        "metric_labels": metric_labels,
        "metric_values": metric_values,
    })


@login_required
def player_create(request):
    form = PlayerForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        player = form.save()
        from core.models import ActivityLog
        ActivityLog.log(request.user, ActivityLog.Action.CREATE,
                        f"Created player: {player.full_name}", "Player", player.pk)
        messages.success(request, _("Player created."))
        return redirect("players:detail", pk=player.pk)
    return render(request, "players/form.html", {"form": form, "title": _("New Player")})


@login_required
def player_edit(request, pk):
    player = get_object_or_404(Player, pk=pk)
    form = PlayerForm(request.POST or None, request.FILES or None, instance=player)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, _("Player updated."))
        return redirect("players:detail", pk=pk)
    return render(request, "players/form.html", {"form": form, "title": _("Edit Player"), "player": player})


@login_required
@require_POST
def player_delete(request, pk):
    player = get_object_or_404(Player, pk=pk)
    if not request.user.can_manage:
        messages.error(request, _("Permission denied."))
        return redirect("players:detail", pk=pk)
    player.delete()
    messages.success(request, _("Player deleted."))
    return redirect("players:list")


@login_required
def add_stats(request, pk):
    player = get_object_or_404(Player, pk=pk)
    form = PlayerStatsForm(request.POST or None, initial={"player": player})
    if request.method == "POST" and form.is_valid():
        stats = form.save(commit=False)
        stats.player = player
        stats.save()
        messages.success(request, _("Stats saved."))
        return redirect("players:detail", pk=pk)
    return render(request, "players/stats_form.html", {"form": form, "player": player})


@login_required
def add_injury(request, pk):
    player = get_object_or_404(Player, pk=pk)
    form = InjuryLogForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        injury = form.save(commit=False)
        injury.player = player
        injury.recorded_by = request.user
        injury.save()
        # Update player status
        player.status = Player.Status.INJURED
        player.save(update_fields=["status"])
        messages.success(request, _("Injury logged."))
        return redirect("players:detail", pk=pk)
    return render(request, "players/injury_form.html", {"form": form, "player": player})


@login_required
def add_metric(request, pk):
    player = get_object_or_404(Player, pk=pk)
    form = PerformanceMetricForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        metric = form.save(commit=False)
        metric.player = player
        metric.recorded_by = request.user
        metric.save()
        messages.success(request, _("Metric recorded."))
        if request.htmx:
            from django.template.loader import render_to_string
            from django.http import HttpResponse
            html = render_to_string("players/partials/metric_row.html", {"metric": metric})
            return HttpResponse(html)
        return redirect("players:detail", pk=pk)
    return render(request, "players/metric_form.html", {"form": form, "player": player})


@login_required
def export_players_excel(request):
    """Export player list to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Players"

    headers = ["ID", "Name", "Position", "Nationality", "Age", "Status", "Sport", "Market Value"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="059669")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    players = Player.objects.all()
    for row, player in enumerate(players, 2):
        ws.cell(row=row, column=1, value=player.pk)
        ws.cell(row=row, column=2, value=player.full_name)
        ws.cell(row=row, column=3, value=player.position)
        ws.cell(row=row, column=4, value=player.nationality)
        ws.cell(row=row, column=5, value=player.age)
        ws.cell(row=row, column=6, value=player.get_status_display())
        ws.cell(row=row, column=7, value=player.get_sport_display())
        ws.cell(row=row, column=8, value=float(player.market_value) if player.market_value else "")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=players.xlsx"
    wb.save(response)
    return response


@login_required
def player_pdf(request, pk):
    player = get_object_or_404(Player, pk=pk)
    return player_profile_pdf(player)
